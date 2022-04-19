### ΕΙΣΑΓΩΓΗ ΔΙΑΦΟΡΩΝ MODULE
import tkinter  
import tkinter.ttk
from tkinter import filedialog
import sys
from sys import *
import openpyxl
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from configparser import ConfigParser
import os
import os.path
###########################ΜΕΤΑΒΛΗΤΕΣ ΠΡΟΓΡΑΜΜΑΤΟΣ####################################################
global WILDCARD  # ΔΗΛΩΝΕΙ ΠΟΣΑ ΜΑΤΣ ΠΡΕΠΕΙ ΝΑ ΠΑΙΞΕΙ Ο ΠΑΙΧΤΗΣ ΓΙΑ ΝΑ ΠΑΡΕΙ ΜΠΑΛΑΝΤΕΡ
WILDCARD = 3
global MAX_ACTIVE_CHALLENGES  # ΔΗΛΩΝΕΙ ΤΟΝ ΜΕΓΙΣΤΟ ΑΡΙΘΜΟ ΠΡΟΚΛΗΣΕΩΝ ΠΟΥ ΜΠΟΡΕΙ ΝΑ EXEI ENERΓΕΣ  ΕΝΑΣ ΠΑΙΧΤΗΣ
MAX_ACTIVE_CHALLENGES =2
global MAX_RANKING_CHALLENGE  # ΔΗΛΩΝΕΙ ΤΗΝ ΜΕΓΙΣΤΗ ΑΠΟΣΤΑΣΗ(ΔΙΑΦΟΡΑ ΚΑΤΑΤΑΞΗΣ) ΠΟΥ ΠΡΕΠΕΙ ΝΑ ΕΧΟΥΝΕ ΟΙ ΠΑΙΚΤΕΣ ΓΙΑ ΝΑ ΣΤΕΙΛΟΥΝΕ ΠΡΟΚΛΗΣΗ
MAX_RANKING_CHALLENGE = 5
global POINT_SET_MATCH # ΟΡΙΖΕΙ ΤΟΝ ΜΕΓΙΣΤΟ ΑΡΙΘΜΟ ΣΕΤ ΠΟΥ ΠΡΕΠΕΙ ΝΑ ΠΑΡΕΙ ΕΝΑΣ ΠΑΙΚΤΗΣ ΓΙΑ ΝΑ ΚΕΡΔΙΣΕΙ ΤΟ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
POINT_SET_MATCH = 3
global filename_ranking # ΔΗΛΩΝΕΙ ΤΟ ΑΡΧΕΙΟ ΕXCEL ΠΟΥ ΣΩΖΟΝΤΑΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ
filename_ranking = os.path.abspath("ranking.xlsx") # ΤΟ ΑΡΧΕΙΟ ΜΕ ΤΗΝ ΚΑΤΑΤΑΞΗΣ
filename_conf = os.path.abspath("config.ini") # ΤΟ ΑΡΧΕΙΟ ΡΥΘΜΙΣΕΩΝ
#########################################################################################################
pl_data = [] #ΛΙΣΤΑ ΜΕ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ
ch_out = [] #ΛΙΣΤΑ ΜΕ ΤΙΣ ΜΗ ΕΓΚΥΡΕΣ ΠΡΟΚΛΗΣΕΙΣ
ch_valid = [] # ΛΙΣΤΑ ΜΕ ΤΙΣ ΕΓΚΥΡΕΣ ΠΡΟΚΛΗΣΕΙΣ
###########################ΟΡΙΣΜΟΣ ΚΛΑΣΗΣ ΠΑΙΚΤΗΣ########################################################
class Player:
    def __init__(self,name,surname,age,m_pl,m_won,m_lost,sets_pl,sets_won,sets_lost,wildcard,active_ch_counter,active_ch):
        self.name = name    #ΟΝΟΜΑ
        self.surname = surname # ΕΠΩΝΥΜΟ
        self.age = age #ΗΛΙΚΙΑ
        self.m_pl = m_pl # ΜΑΤΣ ΣΥΝΟΛΟ
        self.m_won = m_won # ΚΕΡΔΙΣΕΝΑ ΜΑΤΣ
        self.m_lost = m_lost # ΧΑΜΕΝΑ ΜΑΤΣ
        self.sets_pl = sets_pl # ΣΕΤ ΣΥΝΟΛΟ
        self.sets_won = sets_won # ΚΕΡΔΙΣΕΝΑ ΣΕΤ
        self.sets_lost = sets_lost # ΧΑΜΕΝΑ ΣΕΤ
        self.wildcard = wildcard # ΤΟ ΑΝ ΕΧΕΙ ΜΠΑΛΑΝΤΕΡ
        self.active_ch_counter = active_ch_counter # ΤΟ ΠΟΣEΣ ΠΡΟΚΛΗΣΕΙΣ ΕΧΕΙ ΣΥΝΟΛΙΚΑ
        self.active_ch = active_ch # ΤΟ ΕΑΝ ΜΠΟΡΕΙ ΝΑ ΔΕΚΤΕΙ Η ΝΑ ΔΩΣΕΙ ΠΡΟΚΛΗΣΗ Η ΝΑ ΔΕΚΤΕΙ ΠΡΟΚΛΗΣΗ
        return
    def increase_m_pl(self):
        self.m_pl = self.m_pl+1
        return
    def increase_m_won(self):
        self.m_won = self.m_won +1 
        return
    def increase_m_lost(self):
        self.m_lost = self.m_lost +1
        return
    def increase_sets_pl(self):
        self.sets_pl = self.sets_won +self.sets_lost
        return
    def increase_sets_won(self,number):
        self.sets_won += number
        return
    def increase_sets_lost(self,number):
        self.sets_lost += number
        return
    def change_wild_state(self):
        if(self.wildcard == False):
            self.wildcard = True
        if(self.wildcard == True):
            self.wildcard = False
        return
    def check_wild_state(self):
        if(m_pl>=WILDCARD and m_pl%WILDCARD==0):
            change_wild_state(self)
        return
    def increase_active_ch_counter(self):
        self.active_ch_counter+=1
        return
    def decrease_active_ch_counter(self):
        self.active_ch_counter-=1
        return
    def check_active_challenges(self):
        if(self.active_ch_counter>MAX_ACTIVE_CHALLENGES):
            self.active_ch == False
        return
    def __str__(self):
        return f"Ονομα\t\t{self.name}\t\tΕπώνυμο\t\t{self.surname}\t\tΗλικία\t\t{self.age}"
            
    def __stats__(self):
        return f"{self.name}\t\t{self.surname}\tΣύνολο\t{self.m_pl}\tΝίκες\t{self.m_won}\tΗττες\t{self.m_lost}\tΣύνολο σετ\t{self.sets_pl}\tΚερδισμένα σετ\t{self.sets_won}\tΧαμένα σετ\t{self.sets_lost}"
 ################ΟΡΙΣΜΟΣ ΚΛΑΣΗ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###################
class Challenge_Match:
    def __init__(self,t_rank,g_rank,t_name,g_name,t_surname,g_surname):
        self.t_rank = t_rank    #ΚΑΤΑΤΑΞΗ ΠΑΙΚΤΗ ΣΤΟΧΟΥ
        self.g_rank = g_rank    #ΚΑΤΑΤΑΞΗ ΠΑΙΚΤΗ ΠΟΥ ΠΡΟΚΑΛΕΣΕ
        self.t_name = t_name    # ΟΝΟΜΑ ΣΤΟΧΟΥ
        self.g_name = g_name    # ΟΝΟΜΑ ΠΑΙΚΤΗ ΠΟΥ ΠΡΟΚΑΛΕΣΕ
        self.t_surname = t_surname  # ΕΠΙΘΕΤΟ ΠΑΙΚΤΗ ΣΤΟΧΟΥ
        self.g_surname = g_surname  # ΕΠΙΘΕΤΟ ΠΑΙΚΤΗ ΠΟΥ ΠΡΟΚΑΛΕΣΕ
        return
    def __str__(self): ## ΕΚΤΥΠΩΣΗ ΤΗΣ ΚΛΑΣΗΣ ΓΙΑ ΤΑ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
        return f"Κατάταξη\t{(self.t_rank+1)}\t{self.t_name}\t{self.t_surname}\t\tVS\t\tΚατάταξη\t{self.g_rank+1}\t{self.g_name}\t{self.g_surname}" 

        
def make_player(list_a,name,surname,age): #ΔΗΜΙΟΥΡΓΕΙ ΤΟ ΑΝΤΙΚΕΙΜΕΝΟ ΤΗΣ ΚΛΑΣΗΣ ΠΑΙΚΤΗ ΚΑΙ ΤΟ ΑΠΟΘΗΚΕΥΕΙ ΣΕ ΛΙΣΤΑ
    list_a.append(Player(name,surname,age,m_pl=0,m_won=0,m_lost=0,sets_pl=0,sets_won=0,sets_lost=0,wildcard=True,active_ch_counter=0,active_ch=True))
    return  
def make_challenge(list_a,t_rank,g_rank,t_name,g_name,t_surname,g_surname):  #ΔΗΜΙΟΥΡΓΕΙ ΤΟ ΑΝΤΙΚΕΙΜΕΝΟ ΤΗΣ ΚΛΑΣΗΣ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΚΑΙ ΤΟ ΑΠΟΘΗΚΕΥΕΙ ΣΕ ΛΙΣΤΑ
    list_a.append(Challenge_Match(t_rank,g_rank,t_name,g_name,t_surname,g_surname))
    return    
def delete_object(list_a,index): #ΣΒΗΝΕΙ ΕΝΑ ΑΝΤΙΚΕΙΜΕΝΟ ΑΠΟ ΛΙΣΤΑ
    del list_a[index]
    return
def delete_ch_match(list_a,ranking):
    for i in range(len(list_a)):
        if(list[i].t_rank == ranking or list[i].g_rank):
            del list_a[i]
        return
###################ΛΟΓΙΚΗ ΠΡΟΚΛΗΣΕΩΝ##############################################
###TRUE ΕΓΚΥΡΗ ΠΡΟΚΛΗΣΗ                                                          #
###FALSE ΜΗ ΕΓΚΥΡΗ                                                               #
### Η ΚΥΡΙΑ ΛΟΓΙΚΗ ΓΙΑ ΤΟ ΕΑΝ ΜΠΟΡΕΙ ΑΝ ΔΟΘΕΙ ΜΙΑ ΠΡΟΚΛΗΣΗ                       #
##################################################################################
def has_wildcard(t_rank,g_rank,list_a):# ΕΑΝ ΕΧΕΙ ΜΠΑΛΑΝΤΕΡ
    if(t_rank<g_rank and list_a[g_rank].active_ch == True and list_a[g_rank].wildcard==True and list_a[t_rank].active_ch==True):
        return True       
    else:
        return False
def no_wildcard(t_rank,g_rank,list_a):# ΕΑΝ ΔΕΝ ΕΧΕΙ ΜΠΑΛΑΝΤΕΡ
    if(t_rank<g_rank and abs(g_rank - t_rank)<=MAX_RANKING_CHALLENGE and list_a[g_rank].active_ch == True and list_a[t_rank].active_ch == True):
        return True 
    else:
        return False
def duplicate(t_rank,g_rank,list_b):# ΕΛΕΓΧΕΙ ΓΙΑ ΠΡΟΚΛΗΣΗ ΠΟΥ ΕΧΕΙ ΗΔΗ ΔΟΘΕΙ
    if(len(list_b)>0 and list_b[t_rank].t_rank == t_rank and list_b[g_rank].g_rank == g_rank):
        return False
    else:return True
def logic(t_rank,g_rank,list_a,list_b):# ΚΥΡΙΑ ΣΥΝΑΡΤΗΣΗ
    if((has_wildcard(t_rank,g_rank,list_a) == True and duplicate(t_rank,g_rank,list_b)==True) or (no_wildcard(t_rank,g_rank,list_a) == True and duplicate(t_rank,g_rank,list_b) == True)):
        return True
    else: 
        return False
 
def swap(list_a,winner,loser):# Η ΣΥΝΑΡΤΗΣΗ ΑΥΤΗ ΔΕΧΕΤΕ ΜΙΑ ΛΙΣΤΑ ΚΑΙ ΔΥΟ ΑΡΙΘΜΟΥΣ ΚΑΤΑΤΑΞΗΣ ΚΑΙ ΕΠΙΣΤΡΕΦΕΙ ΤΟΝ ΠΙΝΑΚΑ ΜΕ ΒΑΣΗ ΤΗΝ ΑΛΛΑΓΗ ΣΤΗΝ ΚΑΤΑΤΑΞΗ
    base = list_a[loser] # ΑΠΟΘΗΚΕΥΟΥΜΕ ΤΟΝ ΠΑΙΚΤΗ ΠΟΥ ΗΤΤΗΘΗΚΕ 
    list_a[loser] = list_a[winner] # ΣΤΗΝ ΘΕΣΗ ΤΟΥ ΗΤΤΗΜΕΝΟΥ ΠΕΡΝΑΕΙ Ο ΝΙΚΗΤΗΣ
    TEMP = []# ΤΟΠΙΚΕΣ ΜΕΤΑΒΛΗΤΕΣ ΠΟΥ ΣΩΖΟΥΝ ΤΑ ΣΤΟΙΧΕΙΑ ΤΗΣ ΛΙΣΤΑΣ
    temp = []
    for i in range(loser+1,len(list_a)):# ΕΠΑΝΑΛΗΨΗ ΑΠΟ ΤΗΝ ΘΕΣΗ ΤΟΥ ΗΤΤΗΜΕΝΟΥ ΑΥΞΗΜΕΝΗ ΚΑΤΑ ΜΙΑ ΘΕΣΗ ΕΩΣ ΤΟ ΤΕΛΟΣ ΤΙΣ ΛΙΣΤΑΣ
        if(i == loser+1): # ΕΑΝ ΤΟ Ι ΕΙΝΑΙ ΙΣΟ ΜΕ ΤΗΝ ΘΕΣΗ ΜΕΤΑ ΤΟΝ ΧΑΜΕΝΟ
           TEMP = list_a[i]
           temp = list_a[i]
           list_a[i] = base # ΠΕΡΝΑΜΕ ΤΟΝ ΧΑΜΕΝΟ ΜΙΑ ΘΕΣΗ ΠΙΟ ΚΑΤΩ
        if(i%2==0 and i !=loser+1): # ΤΟ ΙΔΙΟ ΚΑΝΟΥΜΕ ΚΑΙ ΓΙΑ ΚΑΘΕ ΑΛΛΗ ΘΕΣΗ
            TEMP = list_a[i]
            list_a[i] = temp
        if(i%2!=0 and i != loser+1):
            temp = list_a[i]
            list_a[i]=TEMP
    return
####################################################################
### ΜΕΝΟΥ###########################################################
####################################################################
###ΑΥΤΗ Η ΚΛΑΣΗ ΕΙΝΑΙ ΤΟ ΚΥΡΙΟ ΜΕΝΟΥ###
class MAIN(tkinter.Tk):
    def __init__(self):
        super().__init__()
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.geometry(f"{self.screen_width}x{self.screen_height}")
        self.configure(background='silver')
        self.label = tkinter.Label(text="Κατατάξη",bg='dark gray',fg = 'red')
        self.label.pack(anchor=tkinter.N,side=tkinter.TOP)
        self.text = tkinter.Text(bg='dark gray',fg = 'red',height=40,width = 180,relief=tkinter.RAISED)
        self.text.pack(anchor=tkinter.N,side=tkinter.TOP)
        self.label_right = tkinter.Label(text="Ενεργα Μάτς Πρόκλησης",bg='dark gray',fg = 'red')
        self.label_right.pack(anchor=tkinter.NE,side = tkinter.TOP)
        self.right_text = tkinter.Text(bg='dark gray',fg='red',height = 40,width = 120,relief=tkinter.RAISED)
        self.right_text.pack(anchor=tkinter.NE,side = tkinter.RIGHT)
        self.label_left = tkinter.Label(text="Εκρρεμή Μάτς Πρόκλησης",bg='dark gray',fg = 'red')
        self.label_left.pack(anchor=tkinter.NW,side = tkinter.TOP)
        self.left_text = tkinter.Text(bg = 'dark gray',fg = 'red',height = 40,width=120,relief=tkinter.RAISED)
        self.left_text.pack(anchor=tkinter.NW,side = tkinter.LEFT)
        self.terminate = tkinter.Button(self,height=1,width=10,text = "Εξοδος",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System",12),highlightcolor="black",command=self.terminate)
        self.terminate.pack(anchor = tkinter.SE,side=tkinter.RIGHT)
        self.settings_button = tkinter.Button(self,height=1,width=10,text = "Ρυθμίσεις",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",bg="silver",font=("System",12),highlightcolor="black",command=self.settings)
        self.settings_button.pack(anchor = tkinter.SW,side=tkinter.BOTTOM)
        self.print_ranking_button = tkinter.Button(self,height=1,width=20,text ="Εκτύπωση Κατάταξης",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System",12),highlightcolor="black",command=self.print_ranking)
        self.print_ranking_button.pack(anchor = tkinter.S,side=tkinter.BOTTOM)
        self.print_stats = tkinter.Button(self,height=1,width=35,text ="Εκτύπωση Κατάταξης Με Στατιστικά",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System",12),highlightcolor="black",command=self.print_ranking_stats)
        self.print_stats.pack(anchor = tkinter.S,side=tkinter.BOTTOM)
        self.print_out = tkinter.Button(self,height=1,width=35,text ="Εκτύπωση Εκρρεμών Μάτς Πρόκλησης",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System",12),highlightcolor="black",command=self.print_out_ch_matches)
        self.print_out.pack(anchor = tkinter.S,side=tkinter.BOTTOM)
        self.print_valid = tkinter.Button(self,height=1,width=35,text ="Εκτύπωση Ενεργών Μάτς Πρόκλησης",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System",12),highlightcolor="black",command=self.print_valid_ch_matches)
        self.print_valid.pack(anchor = tkinter.S,side=tkinter.BOTTOM)
        self.man_out = tkinter.Button(self,height=1,width=35,text ="Διαχείριση Εκρεμών Μάτς Πρόκλησης",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System",12),highlightcolor="black",command=self.ch_out_man)
        self.man_out.pack(anchor = tkinter.S,side=tkinter.BOTTOM)
        self.man_valid = tkinter.Button(self,height=1,width=35,text ="Ενημέρωση Ενεργών Μάτς Πρόκλησης ",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System",12),highlightcolor="black",command=self.ch_valid_man)
        self.man_valid.pack(anchor = tkinter.S,side=tkinter.BOTTOM)
        self.ch_new_match = tkinter.Button(self,height=1,width=20,text ="Νέο Μάτς Πρόκλησης",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System",12),highlightcolor="black",command=self.new_ch_match)
        self.ch_new_match.pack(anchor = tkinter.S,side=tkinter.BOTTOM)
        self.del_pl = tkinter.Button(self,height=1,width=20,text ="Διαγραφή Παίχτη",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System",12),highlightcolor="black",command=self.del_player)
        self.del_pl.pack(anchor = tkinter.S,side=tkinter.BOTTOM)
        self.new_pl = tkinter.Button(self,height=1,width=20,text ="Εγγραφή Νέου Παίχτη",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,bg="silver",fg="black",font=("System",12),highlightcolor="black",command=self.add_player)
        self.new_pl.pack(anchor = tkinter.S,side=tkinter.BOTTOM)
    
    ####ΚΟΥΜΠΙ ΤΕΡΜΑΤΙΣΜΟΥ####
    def terminate(self):
        global filename_player_data
        global filename_out_ch
        global filename_valid_ch
        write_excel(filename_ranking,pl_data,ch_out,ch_valid)
        root.destroy()
        
   
    ###ΤΥΠΩΝΕΙ ΤΗΝ ΒΑΣΙΚΗ ΚΑΤΑΤΑΞΗ###
    def print_ranking(self): 
        self.text.configure(state = 'normal')
        self.text.delete('1.0','end')
        for index in range(len(pl_data)):
            self.text.insert(tkinter.INSERT,str(index+1)+'\t')
            self.text.insert(tkinter.INSERT,pl_data[index].__str__()+'\n')
        self.text.configure(state = 'disabled')
    
    ###TYΠΩΝΕΙ ΤΗΝ ΚΑΤΑΤΑΞΗ ΜΕ ΣΤΑΤΙΣΤΙΚΑ###
    def print_ranking_stats(self):
        self.text.configure(state = 'normal')
        self.text.delete('1.0','end')
        for index in range(len(pl_data)):
            self.text.insert(tkinter.INSERT,str(index+1)+'\t')
            self.text.insert(tkinter.INSERT,pl_data[index].__stats__()+'\n')
        self.text.configure(state = 'disabled')

    ###ΤΥΠΩΝΕΙ ΤΑ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΠΟΥ ΕΙΝΑΙ ΕΚΡΕΜΗ###
    def print_out_ch_matches(self):
        self.left_text.configure(state = 'normal')
        self.left_text.delete('1.0','end')
        for i in range(len(ch_out)):
             self.left_text.insert(tkinter.INSERT,"Πρόκληση"+'\t'+'\t'+str(i+1)+'η'+'\t' + ch_out[i].__str__()+'\n')
        self.left_text.configure(state = 'disabled')

    ###ΤΥΠΩΝΕΙ ΤΑ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΠΟΥ ΕΙΝΑΙ ΕΝΕΡΓΑ###
    def print_valid_ch_matches(self):
        self.right_text.configure(state='normal')
        self.right_text.delete('1.0','end')
        for i in range(len(ch_valid)):
            self.right_text.insert(tkinter.INSERT,"Πρόκληση"+'\t'+str(i+1)+'η'+'\t'+ch_valid[i].__str__()+'\n')
        self.right_text.configure(state = 'disabled')
    ###ΔΙΑΧΕΙΡΙΣΗ ΕΚΚΡΕΜΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def ch_out_man(self):
        top = OUTSTANDING_CHALLENGE_MATCH(self)
        top.grab_set()
    ###ΕΝΗΜΕΡΩΣΗ ΕΝΕΡΓΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def ch_valid_man(self):
        top = BRIEF_CHALLENGE_MATCH(self)
        top.grab_set()
    ###ΝΕΟ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def new_ch_match(self):
        top = NEW_CHALLENGE_MATCH(self)
        top.grab_set()
    ###ΔΙΑΓΡΑΦΗ ΠΑΙΚΤΗ###
    def del_player(self):
        top = DEL_PLAYER(self)
        top.grab_set()
    ###ΕΓΡΑΦΗ ΝΕΟΥ ΠΑΙΧΤΗ###
    def add_player(self):
        top = ADD_PLAYER(self)
        top.grab_set()
    ###ΡΥΘΜΙΣΕΙΣ###
    def settings(self):
        top = SETTINGS(self)
        top.grab_set()

###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΑΠΟΔΟΧΗ ΕΙΣΟΔΟΥ###

class CONFIRMED(tkinter.Toplevel): 
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//10}x{self.screen_height//10}+{self.screen_width//2}+{self.screen_height//2}")
        tkinter.Label(self,bg="silver",text="Αποδοχή",fg="green",font=("System",12)).pack()
        tkinter.Button(self,text='Επιστροφή',activebackground="white",activeforeground="red",fg="black",font=("System",10),highlightcolor="black",bg="silver",padx=10,pady=10,command=self.destroy).pack(anchor=tkinter.S,side=tkinter.BOTTOM)

###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΑΠΟΡΡΙΨΗ ΕΙΣΟΔΟΥ###
class FAILED(tkinter.Toplevel): 
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//10}x{self.screen_height//10}+{self.screen_width//2}+{self.screen_height//2}")
        tkinter.Label(self,bg="silver",text="Απόρριψη",fg="red",font=("System",12)).pack()
        tkinter.Button(self,text='Επιστροφή',activebackground="white",activeforeground="red",bg="silver",fg="black",font=("System",10),highlightcolor="black",padx=10,pady=10,command=self.destroy).pack(anchor=tkinter.S,side=tkinter.BOTTOM)

###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΕΓΓΡΑΦΗ ΝΕΟΥ ΠΑΙΚΤΗ###
class ADD_PLAYER(tkinter.Toplevel):
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//6}x{self.screen_height//6}+{self.screen_width//2}+{self.screen_height//2}")
        self.name = tkinter.StringVar()
        self.surname = tkinter.StringVar()
        self.age = tkinter.IntVar(value=18)
        tkinter.Label(self,text="Ονομα",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="silver",font=("System",12,"bold"),textvariable=self.name).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,text="Επώνυμο",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="silver",font=("System",12,"bold"),textvariable=self.surname).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,text="Ηλικία",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="silver",font=("System",12,"bold"),textvariable=self.age).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,height=1,width=30,text = "Επιστροφή",activebackground="white",bg="silver",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,height=1,width=30,text = "Εγγραφή",activebackground="white",bg="silver",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
###ΕΠΙΒΕΒΑΙΩΣΗ ΠΡΟΣΘΗΚΗΣ ΠΑΙΚΤΗ###
    def confirm(self):
        try:
                name=self.name.get()
                surname=self.surname.get()
                age=self.age.get()
        except:
                top = FAILED(self)
                top.grab_set()
                
        if(len(name)>=1 and len(surname)>=1 and age>=18):
            make_player(pl_data,name,surname,age) #ΠΡΟΣΘΕΤΕΙ ΤΟΝ ΠΑΙΧΤΗ ΕΑN 18+
            root.print_ranking()
            top = CONFIRMED(self)
            top.grab_set()
        
        

###ΚΛΑΣΗ ΥΠΟ ΜΕΝΟΥ ΔΙΑΓΡΑΦΗ ΠΑΙΚΤΗ####
class DEL_PLAYER(tkinter.Toplevel):
    def __init__(self,parent): 
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//6}x{self.screen_height//6}+{self.screen_width//2}+{self.screen_height//2}")
        self.variable_del = tkinter.IntVar()
        tkinter.Label(self,text="Αριθμός Κατάταξης",bg="silver",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,fg="black",bg="silver",font=("System",12,"bold"),textvariable=self.variable_del).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,height=1,width=30,bg="silver",text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,height=1,width=30,bg="silver",text = "Διαγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm_del).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
###EΠΙΒΕΒΑΙΩΣΗ ΔΙΑΓΡΑΦΗ ΠΑΙΚΤΗ###
    def confirm_del(self):
        try:
            var_del_player = self.variable_del.get()-1
        except:
            top = FAILED(self)
            top.grab_set()
               
        if(var_del_player >=0 and var_del_player<=(len(pl_data))):         
            delete_ch_match(ch_out,var_del_player)
            delete_ch_match(ch_valid,var_del_player)
            delete_object(pl_data,var_del_player)
            root.print_ranking()
            root.print_out_ch_matches()
            root.print_valid_ch_matches()
            top = CONFIRMED(self)
            top.grab_set()
        else:
            top = FAILED(self)
            top.grab_set()
               
###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΡΥΘΜΙΣΕΙΣ ##
class SETTINGS(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//6}x{self.screen_height//6}+{self.screen_width//2}+{self.screen_height//2}")
        self.wildcard = tkinter.IntVar()
        self.max_ch_ranking = tkinter.IntVar()
        self.max_active = tkinter.IntVar() 
        self.max_sets = tkinter.IntVar() 
        tkinter.Label(self,bg="silver",text="Μάτς που παίζονται για Μπαλαντέρ",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.wildcard).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Απόσταση στην Κατάταξη μεταξύ παικτών για προκλήση",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.max_ch_ranking).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Mέγιστος αριθμός Επιτρεπώμενων Ενεργών Προκλήσεων",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.max_active).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Mέγιστος αριθμός σετ που πρέπει να κερδίσει ο παίκτης για να πάρει το παιχνίδι",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.max_sets).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,height=1,width=10,bg="silver",text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,height=1,width=15,bg="silver",text = "Φόρτωση Αρχείου",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.load_from_file).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
        tkinter.Button(self,height=1,width=10,bg="silver",text = "Εγγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.define_settings).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
### ΡΥΘΜΙΣΕΙΣ###
    def define_settings(self):
        global WILDCARD
        global MAX_RANKING_CHALLENGE
        global MAX_ACTIVE_CHALLENGES
        global POINT_SET_MATCH
        try:
            WILDCARD = self.wildcard.get()
            MAX_RANKING_CHALLENGE = self.max_ch_ranking.get()
            MAX_ACTIVE_CHALLENGES = self.max_active.get()
            POINT_SET_MATCH = self.max_sets.get()
            write_config()
            top = CONFIRMED(self)
            top.grab_set()
        except:
            top = FAILED(self)
            top.grab_set()
        return 
    
    def load_from_file(self):
        global filename_ranking
        filename_ranking = tkinter.filedialog.askopenfilename(title='Αρχείο Κατάταξης')
        if(filename_ranking == ""):
            filename_ranking = os.path.abspath("ranking.xlsx")
        open_excel(filename_ranking,pl_data,ch_out,ch_valid)
        return
       
    
###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΝΕΟ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
class NEW_CHALLENGE_MATCH(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//6}x{self.screen_height//6}+{self.screen_width//2}+{self.screen_height//2}")
        self.challenger = tkinter.IntVar()
        self.champion = tkinter.IntVar()
        tkinter.Label(self,bg="silver",text="Κατάταξη Παίκτη που προκαλεί",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.challenger).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Κατάταξη Παίκτη αποδέκτης πρόκλησης",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.champion).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,bg="silver",height=1,width=10,text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,bg="silver",height=1,width=10,text = "Εγγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm_ch_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
    def confirm_ch_match(self):
        try:
            g_rank = self.challenger.get()-1
            t_rank = self.champion.get()-1          
        except:
            top = FAILED(self)
            top.grab_set()
        if( g_rank >=0 and g_rank <= (len(pl_data))  and t_rank >=0 and t_rank <=len(pl_data) and g_rank>t_rank):
           ### ΕΛΕΓΧΕΙ ΑΝ ΟΙ ΑΡΙΘΜΟΙ ΕΙΝΑΙ ΣΤΗΝ ΕΠΙΤΡΕΠΟΜΕΝΗ ΑΠΟΣΤΑΣΗ
            pl_data[g_rank].increase_active_ch_counter()
            pl_data[g_rank].check_active_challenges()
            if(logic(t_rank,g_rank,pl_data,ch_out)==True):
                pl_data[g_rank].change_wild_state()
                make_challenge(ch_out,t_rank,g_rank,pl_data[t_rank].name,pl_data[g_rank].name,pl_data[t_rank].surname,pl_data[g_rank].surname)
                root.print_out_ch_matches()
                top = CONFIRMED(self)
                top.grab_set()
                    
                    
            else:
                top = FAILED(self)
                top.grab_set()
                  
        else:
            top = FAILED(self)
            top.grab_set()          

###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΕΝΗΜΕΡΩΣΗ ΕΝΕΡΓΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
class BRIEF_CHALLENGE_MATCH(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//6}x{self.screen_height//6}+{self.screen_width//2}+{self.screen_height//2}")
        self.serial = tkinter.IntVar()
        self.winner = tkinter.IntVar()
        self.loser = tkinter.IntVar()
        self.sets_winner = tkinter.IntVar()
        self.sets_loser = tkinter.IntVar()
        tkinter.Label(self,bg="silver",text="Αριθμός μάτς κατάταξης προς ενημέρωση",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.serial).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Αριθμός κατάταξης παίχτη που νίκησε",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.winner).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Σετ που πήρε ο Νικητής",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.sets_winner).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Αριθμός κατάταξης παίχτη που ηττήθηκε",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.loser).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Label(self,bg="silver",text="Σετ που πήρε ο Ηττημένος",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.sets_loser).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,bg="silver",height=1,width=10,text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,bg="silver",height=1,width=10,text = "Εγγραφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.brief_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
    def brief_match(self):
        
        serial = self.serial.get()-1
        winner = self.winner.get()-1
        loser = self.loser.get()-1
        sets_winner = self.sets_winner.get()
        sets_loser = self.sets_loser.get()        
        if( sets_winner>=1 and sets_winner<=POINT_SET_MATCH and sets_loser>=1 and sets_loser<=POINT_SET_MATCH and sets_winner>sets_loser and winner>=0 and winner<=len(pl_data) and loser>=0 and loser<=len(pl_data) and serial>=0 and serial<=(len(ch_valid))):#EΑΝ ΟΙ ΑΡΙΘΜΟΙ ΥΠΟΚΕΙΝΤΑΙ ΣΤΟΥΣ ΚΑΝΟΝΕΣ ΓΙΑ ΤΑ ΠΑΙΧΝΔΙΑ
            pl_data[winner].increase_m_pl()
            pl_data[loser].increase_m_pl()
            pl_data[winner].increase_m_won()
            pl_data[loser].increase_m_lost()
            pl_data[winner].increase_sets_won(sets_winner)
            pl_data[winner].increase_sets_lost(sets_loser)
            pl_data[loser].increase_sets_won(sets_loser)
            pl_data[loser].increase_sets_lost(sets_winner)
            pl_data[winner].increase_sets_pl()
            pl_data[loser].increase_sets_pl()
            pl_data[winner].decrease_active_ch_counter()
            pl_data[loser].decrease_active_ch_counter()
            swap(pl_data,winner,loser)
            delete_object(ch_valid,serial)
            root.print_valid_ch_matches()
            root.print_ranking()
            top = CONFIRMED(self)
            top.grab_set() 
        else:
            top = FAILED(self)
            top.grab_set()
                
###ΚΛΑΣΗ ΥΠΟΜΕΝΟΥ ΔΙΑΧΕΙΡΙΣΗ ΕΚΡΕΜΜΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
class OUTSTANDING_CHALLENGE_MATCH(tkinter.Toplevel):
    def __init__(self,parent):
        super().__init__(parent)
        self.title("Πρόγραμμα Κατάταξης Τέννις")
        self.screen_width = self.winfo_screenwidth()
        self.screen_height = self.winfo_screenheight()
        self.configure(background='silver')
        self.geometry(f"{self.screen_width//6}x{self.screen_height//6}+{self.screen_width//2}+{self.screen_height//2}")
        self.match = tkinter.IntVar()
        tkinter.Label(self,bg="silver",text="Αριθμός εκρεμούς μάτς κατάταξης",fg="black",font=("System",12,"bold")).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Entry(self,bg="silver",fg="black",font=("System",12,"bold"),textvariable=self.match).pack(anchor=tkinter.N,side=tkinter.TOP)
        tkinter.Button(self,bg="silver",height=1,width=10,text = "Επιστροφή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.destroy).pack(anchor=tkinter.SW,side=tkinter.LEFT)
        tkinter.Button(self,bg="silver",height=1,width=10,text = "Αποδοχή",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.confirm_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
        tkinter.Button(self,bg="silver",height=1,width=10,text = "Απόρριψη",activebackground="white",activeforeground="red",bd=8,relief=tkinter.RAISED,fg="black",font=("System",12),highlightcolor="black",command=self.deny_match).pack(anchor=tkinter.SW,side=tkinter.RIGHT)
    ###ΑΠΟΔΟΧΗ ΕΚΡΕΜΜΟΥΣ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def confirm_match(self):
        try:
            number = self.match.get() - 1
        except:
            pop_up = FAILED(self)
            pop_up.grab_set()
        if(number>=0 and number <=(len(ch_out))):
            t_rank = ch_out[number].t_rank
            g_rank = ch_out[number].g_rank
            if(pl_data[t_rank].active_ch==True):
                pl_data[t_rank].increase_active_ch_counter()
                pl_data[t_rank].check_active_challenges()
                make_challenge(ch_valid,t_rank,g_rank,pl_data[t_rank].name,pl_data[g_rank].name,pl_data[t_rank].surname,pl_data[g_rank].surname)                
                delete_object(ch_out,number)
                root.print_out_ch_matches()
                root.print_valid_ch_matches()
                pop_up = CONFIRMED(self)
                pop_up.grab_set()
            else:
                    pop_up = FAILED(self)
                    pop_up.grab_set()
        else:
                pop_up = FAILED(self)
                pop_up.grab_set()

    ###ΑΡΝΗΣΗ ΕΚΡΕΜΟΥΣ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ###
    def deny_match(self):
        try:
            number = self.match.get() -1
        except:
            pop_up = FAILED(self)
            pop_up.grab_set()
        if(number>=0 and number<=len(ch_out)):
            pl_data[ch_out[number].g_rank].decrease_active_ch_counter()
            pl_data[ch_out[number].g_rank].check_active_challenges()
            delete_object(ch_out,number)
            root.print_out_ch_matches()
            pop_up = CONFIRMED(self)
            pop_up.grab_set()
        else:
            pop_up = FAILED(self)
            pop_up.grab_set()
               
               
def open_excel(filename,list_a,list_b,list_c):
    workbook = openpyxl.load_workbook(filename)
    workbook.active = 0
    worksheet = workbook.active 
    for row in worksheet.iter_rows(): # ΔΙΑΒΑΖΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ
        for cell in row:
            if 'A' in cell.coordinate:
                name = cell.value
            if 'B' in cell.coordinate:
                surname = cell.value
            if 'C' in cell.coordinate:
                age = cell.value
            if 'D' in cell.coordinate:    
                m_pl = cell.value
            if 'E' in cell.coordinate:        
                m_won = cell.value
            if 'F' in cell.coordinate:
                m_lost = cell.value
            if 'G' in cell.coordinate:
                sets_pl = cell.value
            if 'H' in cell.coordinate:
                sets_won = cell.value
            if 'I' in cell.coordinate:
                sets_lost = cell.value
            if 'J' in cell.coordinate:
                wildcard = cell.value
            if 'K' in cell.coordinate:
                active_ch_counter = cell.value
            if 'L' in cell.coordinate:
                active_ch = cell.value
        list_a.append(Player(name,surname,age,m_pl,m_won,m_lost,sets_pl,sets_won,sets_lost,wildcard,active_ch_counter,active_ch))
    workbook.active = 1
    worksheet = workbook.active
    for row in worksheet.iter_rows(): # ΔΙΑΒΑΖΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΜΗ  ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
        for cell in row:
            if 'A' in cell.coordinate:
                t_rank = cell.value
            if 'B' in cell.coordinate:
                g_rank = cell.value
            if 'C' in cell.coordinate:
                t_name = cell.value
            if 'D' in cell.coordinate:
                g_name = cell.value
            if 'E' in cell.coordinate:
                t_surname = cell.value
            if 'F' in cell.coordinate:
                g_surname = cell.value
        list_b.append(Challenge_Match(t_rank,g_rank,t_name,g_name,t_surname,g_surname))
    workbook.active = 2
    worksheet = workbook.active
    for row in worksheet.iter_rows(): # ΔΙΑΒΑΖΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ
        for cell in row:
            if 'A' in cell.coordinate:
                t_rank = cell.value
            if 'B' in cell.coordinate:
                g_rank = cell.value
            if 'C' in cell.coordinate:
                t_name = cell.value
            if 'D' in cell.coordinate:
                g_name = cell.value
            if 'E' in cell.coordinate:
                t_surname = cell.value
            if 'F' in cell.coordinate:
                g_surname = cell.value
        list_c.append(Challenge_Match(t_rank,g_rank,t_name,g_name,t_surname,g_surname))    
    return 
    
def write_excel(filename,list_a,list_b,list_c):
    workbook = openpyxl.load_workbook(filename,read_only=False,keep_vba=False)
    sheets = workbook.sheetnames
    for i in range(len(sheets)):
        workbook.remove(workbook[sheets[i]])
    worksheet = workbook.create_sheet()
    workbook.active = 0
    worksheet = workbook.active
    temp = []
    
    for i in range(len(list_a)): # ΓΡΑΦΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΠΑΙΚΤΩΝ ΣΤΟ ΕΞΕΛ
        j = 1
        for value in list_a[i].__dict__.values():
            worksheet.cell(i+1,j).value = value
            j+=1
    worksheet = workbook.create_sheet()
    workbook.active = 1
    worksheet = workbook.active
    
    for i in range(len(list_b)): # ΓΡΑΦΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΜΗ ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΕΩΝ ΣΤΟ ΕΞΕΛ
        j = 1
        for value in list_b[i].__dict__.values():
            worksheet.cell(i+1,j).value = value
            j+=1
    worksheet = workbook.create_sheet()
    workbook.active = 2
    worksheet = workbook.active
    
    for i in range(len(list_c)): # ΓΡΑΦΕΙ ΤΑ ΣΤΟΙΧΕΙΑ ΤΩΝ ΕΓΚΥΡΩΝ ΜΑΤΣ ΠΡΟΚΛΗΣΗΣ ΣΤΟ ΕΞΕΛ
        j=1
        for value in list_c[i].__dict__.values():
            worksheet.cell(i+1,j).value = value
            j+=1
    workbook.save(filename)
    return
def write_config():#ΓΡΑΦΕΙ ΤΙΣ ΡΥΘΜΙΣΕΙΣ ΣΕ ΑΡΧΕΙΟ ΙΝΙ
    global WILDCARD
    global MAX_ACTIVE_CHALLENGES
    global MAX_RANKING_CHALLENGE
    global POINT_SET_MATCH
    global filename_player_data
    global filename_out_ch
    global filename_valid_ch
    
    config_object = ConfigParser()
    config_object["settings"] = {"WILDCARD":WILDCARD,"MAX_ACTIVE_CHALLENGES":MAX_ACTIVE_CHALLENGES,"MAX_RANKING_CHALLENGE":MAX_RANKING_CHALLENGE,"POINT_SET_MATCH":POINT_SET_MATCH,"ranking": filename_ranking }
    
    with  open("config.ini",'w') as conf:
        config_object.write(conf)
        conf.close
    return

def read_config():# ΔΙΑΒΑΖΕΙ ΤΙΣ ΡΥΘΜΙΣΕΙΣ ΑΠΟ ΑΡΧΕΙΟ ΙΝΙ
    global WILDCARD
    global MAX_ACTIVE_CHALLENGES
    global MAX_RANKING_CHALLENGE
    global POINT_SET_MATCH
    global filename_ranking
    if(os.path.exists(filename_conf)==True):
        config_object = ConfigParser()
        config_object.read("config.ini")
        settings = config_object["settings"]
        WILDCARD = int(settings["WILDCARD"])
        MAX_ACTIVE_CHALLENGES = int(settings["MAX_ACTIVE_CHALLENGES"])
        MAX_RANKING_CHALLENGE = int(settings["MAX_RANKING_CHALLENGE"])
        POINT_SET_MATCH = int(settings["POINT_SET_MATCH"])
        filename_ranking= settings['ranking']
    else:
        write_config()
        read_config()
    return

def startup():
    read_config()
    open_excel(filename_ranking,pl_data,ch_out,ch_valid)
    return
### MAIN###
 
if __name__ == '__main__':
    startup()
    root = MAIN()
    root.print_ranking()
    root.print_valid_ch_matches()
    root.print_out_ch_matches()
    root.mainloop()